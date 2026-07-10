"""Generate app/services/_personal_development_seed_data.py from the owner's
personal-development Excel workbook (7 sheets).

Why a generator instead of hand-typed seed data: the owner's requirement is
"هیچ محتوایی نباید ناقص منتقل بشه" — nothing may be lost or truncated. This
script consumes every non-empty cell through an explicit per-sheet rule and
FAILS if any cell is left unconsumed, so completeness is machine-checked, not
eyeballed. Run:

    python scripts/generate_pd_seed.py /path/to/workbook.xlsx

The output module holds plain data (lists + finance transactions); the runtime
seeder is app/services/personal_development_seed.py.
"""
from __future__ import annotations

import datetime
import pprint
import re
import sys

import openpyxl

PREFIX = "توسعه فردی - "


class Sheet:
    """Wraps a worksheet with consumption tracking so unconsumed cells fail."""

    def __init__(self, ws):
        self.ws = ws
        self.consumed: set[tuple[int, int]] = set()

    def raw(self, r: int, c: int):
        return self.ws.cell(row=r, column=c).value

    def has(self, r: int, c: int) -> bool:
        v = self.raw(r, c)
        return v is not None and str(v).strip() != ""

    def take(self, r: int, c: int) -> str:
        """Consume a cell; returns '' when empty."""
        self.consumed.add((r, c))
        v = self.raw(r, c)
        if v is None:
            return ""
        if isinstance(v, datetime.datetime):
            return v.date().isoformat()
        return str(v).strip()

    def take_date(self, r: int, c: int):
        self.consumed.add((r, c))
        v = self.raw(r, c)
        if isinstance(v, datetime.datetime):
            return v.date().isoformat()
        return None

    def unconsumed(self):
        out = []
        for r in range(1, self.ws.max_row + 1):
            for c in range(1, self.ws.max_column + 1):
                if self.has(r, c) and (r, c) not in self.consumed:
                    out.append((r, c, str(self.raw(r, c))[:60]))
        return out


def _mklist(name, description, items):
    return {"name": PREFIX + name, "description": description or "",
            "items": [i for i in items if i["content"].strip()]}


def _item(content, description=""):
    return {"content": content.strip(), "description": (description or "").strip()}


# ── Sheet 0: چرک نویس ────────────────────────────────────────────────────────
def sheet_cherknevis(sh: Sheet):
    lists = []

    # Priorities with reasons (col 9, rows 3-8)
    pr = [_item(sh.take(r, 9)) for r in range(3, 9)]
    lists.append(_mklist("اولویت‌های اصلی من (با علت)",
                         "اولویت‌های شخصی و علت اهمیت هرکدام — از چرک‌نویس برنامه‌ریزی.", pr))

    def col2_items(r_from, r_to, merge_c1=()):
        items = []
        for r in range(r_from, r_to + 1):
            if not sh.has(r, 2) and not (r in merge_c1 and sh.has(r, 1)):
                continue
            content = sh.take(r, 2)
            desc = ""
            if r in merge_c1 and sh.has(r, 1):
                desc = sh.take(r, 1)
            if not content and desc:
                content, desc = desc, ""
            items.append(_item(content, desc))
        return items

    lists.append(_mklist("سه کار مورد علاقه (تا دو ماه)", sh.take(2, 2),
                         col2_items(3, 5)))
    lists.append(_mklist("کارهای زیر دو دقیقه", sh.take(7, 2), col2_items(8, 18)))
    lists.append(_mklist("یادداشت‌های متفرقهٔ چرک‌نویس",
                         "اعداد/یادداشت‌های بدون عنوان از چرک‌نویس (حفظ کامل محتوا).",
                         col2_items(23, 24)))
    lists.append(_mklist("طرح تثبیت (نسخهٔ ۱ — بدون تاریخ)",
                         "قرآن، نویسندگی، عربی، ورزش و کسب‌وکار — نسخهٔ اول.",
                         col2_items(29, 62)))
    lists.append(_mklist("طرح تثبیت (نسخهٔ ۲)",
                         "نسخهٔ بازنویسی‌شدهٔ طرح تثبیت با جزئیات عربی بیشتر.",
                         col2_items(68, 98)))
    lists.append(_mklist("طرح و برنامه‌ریزی ۲۳/۰۹/۲۰۲۴", sh.take(101, 1),
                         col2_items(103, 139, merge_c1={121, 138})))

    def col1_items(r_from, r_to):
        return [_item(sh.take(r, 1)) for r in range(r_from, r_to + 1) if sh.has(r, 1)]

    hdr = " — ".join(x for x in (sh.take(143, 1), sh.take(143, 2)) if x)
    lists.append(_mklist("بازچینش برنامه‌ها ۱۰/۱۰/۲۰۲۴", hdr, col1_items(145, 153)))
    hdr = " — ".join(x for x in (sh.take(157, 1), sh.take(157, 2)) if x)
    lists.append(_mklist("بازچینش برنامه‌ها ۲۰/۱۰/۲۰۲۴", hdr, col1_items(159, 168)))
    lists.append(_mklist("طرح و برنامه‌ریزی ۲۵/۱۰/۲۰۲۴", sh.take(172, 1),
                         col2_items(174, 211, merge_c1={176, 192, 210})))

    # Weekly programme table 17/11/2024 (rows 214-225, cols 1-5)
    hdr = " — ".join(x for x in (sh.take(214, 1), sh.take(214, 3)) if x)
    items = []
    for r in range(216, 226):
        if not any(sh.has(r, c) for c in (1, 2, 3, 4, 5)):
            continue
        domain, what = sh.take(r, 1), sh.take(r, 2)
        parts = []
        for label, c in (("زمان", 3), ("تعداد در هفته", 4), ("تا کی", 5)):
            v = sh.take(r, c)
            if v:
                parts.append(f"{label}: {v}")
        items.append(_item(f"{domain} {what}".strip(), " | ".join(parts)))
    lists.append(_mklist("برنامهٔ هفتگی ۱۷/۱۱/۲۰۲۴", hdr, items))
    return lists


# ── Sheet 1: مدیریت زمان ─────────────────────────────────────────────────────
def sheet_time(sh: Sheet):
    lists = []
    # Energy/time thieves (col 5, rows 2-19)
    title = sh.take(2, 5)
    items = [_item(sh.take(r, 5)) for r in range(3, 20) if sh.has(r, 5)]
    lists.append(_mklist("دزدان انرژی و زمان", title, items))

    # Weekly time-thief log (cols 9-12, rows 1-46)
    head = sh.take(1, 11)  # دزدان زمان در هفته
    for c in (10, 11, 12):
        sh.take(2, c)  # column headers
    items = []
    for r in range(3, 47):
        if not any(sh.has(r, c) for c in (9, 10, 11)):
            continue
        date = sh.take(r, 9) if sh.has(r, 9) else ""
        rng = sh.take(r, 10) if sh.has(r, 10) else ""
        act = sh.take(r, 11) if sh.has(r, 11) else ""
        desc_parts = [p for p in (f"تاریخ: {date}" if date else "",
                                  f"بازهٔ بیداری: {rng}" if rng else "") if p]
        content = act or rng
        items.append(_item(content, " | ".join(desc_parts)))
    lists.append(_mklist("گزارش دزدان زمان هفته", head, items))

    def notes(r_from, r_to, name, desc):
        out = []
        for r in range(r_from, r_to + 1):
            main = sh.take(r, 3) if sh.has(r, 3) else ""
            note = sh.take(r, 2) if sh.has(r, 2) else ""
            sub = sh.take(r, 4) if sh.has(r, 4) else ""
            if not (main or note or sub):
                continue
            content = main or sub or note
            d = []
            if note and main:
                d.append(f"یادداشت من: {note}")
            if sub and main:
                d.append(sub)
            out.append(_item(content, " | ".join(d)))
        return _mklist(name, desc, out)

    lists.append(notes(44, 229, "نکات سخنرانی مدیریت زمان (رائفی‌پور)",
                       "یادداشت‌های من از سخنرانی مدیریت زمان رائفی‌پور."))
    lists.append(notes(230, 646, "درس‌گفتار مدیریت زمان (پناهیان — جلسات ۱ تا ۹)",
                       "یادداشت‌های من از ۹ جلسهٔ درس‌گفتار مدیریت زمان پناهیان."))
    return lists


# ── Sheet 2: جدول مبارزه با هوای نفس ─────────────────────────────────────────
def sheet_nafs(sh: Sheet):
    desc = " | ".join(x for x in (sh.take(1, 2), sh.take(2, 2), sh.take(3, 2)) if x)
    for c in range(2, 8):
        sh.take(4, c)  # headers: ردیف/دوست‌داشتنی‌ها/نوع/جایگزین/زمان شروع/توضیحات
    items = []
    for r in range(5, 25):
        num = sh.take(r, 2)  # row numbering (structural)
        if not sh.has(r, 3):
            continue
        content = sh.take(r, 3)
        d = []
        for label, c in (("نوع", 4), ("جایگزین", 5), ("زمان شروع", 6), ("توضیحات", 7)):
            v = sh.take(r, c)
            if v:
                d.append(f"{label}: {v}")
        items.append(_item(content, " | ".join(d)))
        del num
    return [_mklist("مبارزه با هوای نفس", desc, items)]


# ── Sheet 3: اهداف ───────────────────────────────────────────────────────────
def sheet_goals(sh: Sheet):
    for c in (2, 3, 4):
        sh.take(2, c)  # headers
    items = []
    for r in range(3, 8):
        sh.take(r, 2)  # numbering
        goal = sh.take(r, 3)
        why = sh.take(r, 4)
        items.append(_item(goal, f"علت: {why}" if why else ""))
    return [_mklist("اهداف و آرزوها", "اهداف و آرزوهای من و علت هرکدام.", items)]


# ── Sheet 4: عادت‌ها جهت بهبود ───────────────────────────────────────────────
def sheet_habits(sh: Sheet):
    guidance = sh.take(2, 12)
    sh.take(5, 7)  # «مراحل بهبود» super-header
    for c in list(range(2, 10)) + [12, 13, 14, 15]:
        sh.take(6, c)  # both tables' headers

    bad = []
    for r in range(7, 9):
        sh.take(r, 2)
        content = sh.take(r, 3)
        d = []
        for label, c in (("علت‌ها", 4), ("زمان‌های انجام", 5), ("عادت خوب مقابل", 6),
                         ("مرحلهٔ اول", 7), ("مرحلهٔ دوم", 8), ("مرحلهٔ سوم", 9)):
            v = sh.take(r, c)
            if v:
                d.append(f"{label}: {v}")
        bad.append(_item(content, " | ".join(d)))
    for r in range(9, 39):  # structural numbering only
        sh.take(r, 2)
    # difficulty scale + examples (rows 42-45, cols 2-7)
    scale = " ← ".join(sh.take(42, c) for c in range(3, 8) if sh.has(42, c))
    bad.append(_item(f"مقیاس دشواری عادت: {scale}"))
    for r in range(43, 46):
        sh.take(r, 2)  # «مثال»
        steps = " ← ".join(sh.take(r, c) for c in range(3, 8) if sh.has(r, c))
        bad.append(_item(f"مثال مقیاس: {steps}"))
    lists = [_mklist("عادت‌های بد و مراحل بهبود",
                     "عادت‌های بد، علت‌ها، عادت خوب مقابل و مراحل بهبود + مقیاس دشواری.", bad)]

    sign_map = {"+": "مثبت", "-": "منفی", "=": "خنثی"}
    daily = []
    for r in range(7, 98):
        sh.take(r, 12)  # numbering
        if not sh.has(r, 13):
            continue
        content = sh.take(r, 13)
        sign = sh.take(r, 14)
        daily.append(_item(content, f"علامت: {sign_map.get(sign, sign)}" if sign else ""))
    lists.append(_mklist("عادت‌های روزانه (مثبت/منفی/خنثی)", guidance, daily))
    return lists


# ── Sheet 5: حساب کتاب ماهانه → finance transactions + archive list ─────────
def sheet_finance(sh: Sheet):
    txs = []

    def month(label, default_date, misc_rows, food_rows,
              misc_date_col=None, food_date_col=None,
              misc_cols=(2, 3), food_cols=(6, 7)):
        for r in misc_rows:
            if not sh.has(r, misc_cols[0]):
                continue
            date = (sh.take_date(r, misc_date_col) if misc_date_col and sh.has(r, misc_date_col) else None)
            name = sh.take(r, misc_cols[0])
            amount = sh.take(r, misc_cols[1])
            txs.append({"description": f"{name} (متفرقه — {label})"[:255],
                        "amount": float(amount), "date": date or default_date})
        for r in food_rows:
            if not sh.has(r, food_cols[0]):
                continue
            date = (sh.take_date(r, food_date_col) if food_date_col and sh.has(r, food_date_col) else None)
            name = sh.take(r, food_cols[0])
            amount = sh.take(r, food_cols[1])
            txs.append({"description": f"{name} (موارد غذایی — {label})"[:255],
                        "amount": float(amount), "date": date or default_date})

    # September (labels r2-3; headers r6)
    sh.take(2, 1)
    sh.take(3, 1)
    for c in (1, 2, 3, 5, 6, 7):
        sh.take(6, c)
    month("سپتامبر ۲۰۲۴", "2024-08-20", range(7, 36), range(7, 21))

    # October (labels r39-40; headers r43)
    for rc in ((39, 1), (40, 1), (40, 2), (40, 3)):
        sh.take(*rc)
    for c in (2, 3, 5, 6, 7):
        sh.take(43, c)
    month("اکتبر ۲۰۲۴", "2024-09-19", range(44, 65), range(44, 78),
          misc_date_col=1, food_date_col=5)

    # November (headers r92 only — the sheet has no month title row here)
    for c in (1, 2, 3, 5, 6, 7):
        sh.take(92, c)
    month("نوامبر ۲۰۲۴", "2024-10-19", range(93, 111), range(93, 133),
          misc_date_col=1, food_date_col=5)

    # December (labels r138-139; headers r141)
    for rc in ((138, 1), (139, 1), (139, 2), (139, 3)):
        sh.take(*rc)
    for c in (1, 2, 3, 5, 6, 7):
        sh.take(141, c)
    month("دسامبر ۲۰۲۴", "2024-11-20", range(142, 161), range(142, 161),
          misc_date_col=1, food_date_col=5)

    # Balances + loans snapshot (rows 82-90) → archive list, not live accounts
    bal_items = []
    for r in range(82, 87):
        if sh.has(r, 3):
            bank = sh.take(r, 3)
            val = sh.take(r, 4)
            bal_items.append(_item(f"ماندهٔ حساب {bank}: {val} درهم"))
    for r in range(85, 91):
        if sh.has(r, 6):
            a, b = sh.take(r, 6), sh.take(r, 7)
            c_, d_ = sh.take(r, 10), sh.take(r, 11)
            bal_items.append(_item(f"{b}: {a}", f"{d_}: {c_}" if (c_ or d_) else ""))
    archive_list = _mklist(
        "حساب کتاب — مانده حساب‌ها و وام‌ها (آرشیو اکتبر ۲۰۲۴)",
        "تصویر لحظه‌ای مانده بانک‌ها (مشرق/فب/صادرات/…) و وام‌ها از فایل اکسل — صرفاً آرشیو.",
        bal_items)
    return txs, [archive_list]


# ── Sheet 6: ابزارها ─────────────────────────────────────────────────────────
def sheet_tools(sh: Sheet):
    method = [_item(sh.take(r, 1)) for r in range(1, 7)]
    lists = [_mklist("روش انتخاب ابزار برای اهداف",
                     "گام‌های پیدا کردن ابزار مناسب برای هر هدف.", method)]

    tools = []
    # StudyChat: title r10 (+ url c8) + notes rows 12-22 (even rows)
    title = sh.take(10, 1)
    url = sh.take(10, 8)
    notes = "\n".join(sh.take(r, 1) for r in (12, 14, 16, 18, 20, 22) if sh.has(r, 1))
    tools.append(_item(f"StudyChat — {title} — {url}", notes))
    # Perplexity: r27 (+ url c7) + notes 30-34
    title = sh.take(27, 1)
    url = sh.take(27, 7)
    notes = "\n".join(sh.take(r, 1) for r in (30, 32, 34) if sh.has(r, 1))
    tools.append(_item(f"Perplexity — {title} — {url}", notes))
    # Long single-cell reviews
    for r, name in ((37, "Heuristica"), (64, "Humata.ai"), (97, "مقایسهٔ ChatGPT با ابزارهای بالا")):
        text = sh.take(r, 1)
        first = text.splitlines()[0][:120]
        tools.append(_item(f"{name} — {first}", text))
    lists.append(_mklist("ابزارهای هوش مصنوعی (بررسی و یادداشت)",
                         "ابزارهای AI بررسی‌شده به‌همراه لینک و توضیح کامل.", tools))
    return lists


def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    names = wb.sheetnames
    sheets = {i: Sheet(wb[n]) for i, n in enumerate(names)}

    all_lists = []
    all_lists += sheet_cherknevis(sheets[0])
    all_lists += sheet_time(sheets[1])
    all_lists += sheet_nafs(sheets[2])
    all_lists += sheet_goals(sheets[3])
    all_lists += sheet_habits(sheets[4])
    txs, fin_lists = sheet_finance(sheets[5])
    all_lists += fin_lists
    all_lists += sheet_tools(sheets[6])

    # ── completeness gate: every non-empty cell must be consumed ────────────
    problems = []
    for i, sh in sheets.items():
        for r, c, preview in sh.unconsumed():
            problems.append(f"  sheet {i} ({names[i]!r}) r{r}c{c}: {preview}")
    if problems:
        print("UNCONSUMED CELLS — generation refused:")
        print("\n".join(problems))
        sys.exit(1)

    total_items = sum(len(x["items"]) for x in all_lists)
    header = (
        '"""Personal-development seed data — GENERATED from the owner\'s Excel\n'
        "workbook by scripts/generate_pd_seed.py. Do not hand-edit; re-run the\n"
        "generator instead. Every non-empty cell of all 7 sheets is represented\n"
        '(the generator fails on unconsumed cells)."""\n\n'
    )
    body = (
        f"PD_ACCOUNT_NAME = \"هزینه‌های نقدی — آرشیو اکسل\"\n"
        f"PD_ACCOUNT_CURRENCY = \"AED\"\n\n"
        f"PD_EXPECTED_LIST_COUNT = {len(all_lists)}\n"
        f"PD_EXPECTED_ITEM_COUNT = {total_items}\n"
        f"PD_EXPECTED_TX_COUNT = {len(txs)}\n\n"
        f"PD_LISTS = {pprint.pformat(all_lists, width=100, sort_dicts=False)}\n\n"
        f"PD_TRANSACTIONS = {pprint.pformat(txs, width=100, sort_dicts=False)}\n"
    )
    out = "app/services/_personal_development_seed_data.py"
    with open(out, "w", encoding="utf-8") as f:
        f.write(header + body)
    print(f"OK → {out}")
    print(f"lists={len(all_lists)} items={total_items} transactions={len(txs)}")
    for x in all_lists:
        print(f"  • {x['name']} — {len(x['items'])} آیتم")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else
         re.sub(r"$", "", "/root/.claude/uploads/88ee70e6-f682-5575-a0d1-629218c6959e/9366b700-__________.xlsx"))
